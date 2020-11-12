import json
import argparse
from pathlib import Path
from openpyxl.utils import FORMULAE
from openpyxl.styles import Alignment
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import set_style_cards, set_style_creators
from classes import Card, Result, Creator
from config import LABELS, TIME_LABELS



def parse_args(input_args=None):
    parser = argparse.ArgumentParser(description="Создает файл excel с необходимыми данными.")
    parser.add_argument("--input_path", help='Входной путь к файлу', required=True)
    parser.add_argument("--output_path", help='Результат', required=True)
    args = parser.parse_known_args(input_args)[0]
    return args


def json_reader(filename, data=None):
    '''Открывает json-файл и конвертирует его в объект python'''
    with open(filename, 'r') as file:
        data = json.load(file)
    return data


def get_cards(data,cards = []):
    '''Создает объект список объектов из json-файла Card'''

    for card in data['cards']:
        card = Card(id=card['_id'],
                    title=card['title'],
                    creator=card['userId'],
                    labels=(", ".join(card['labelIds'])))
        cards.append(card)
    return cards

def change_id_to_name(data,cards,users={},labels={}):
    '''Изменяет все id объектов на имена'''
    for user in data['users']:
        users[user['_id']] = user['username']
    for label in data['labels']:
        labels[label['_id']] = label['name']
    for card in cards:
        lab = card.labels
        card.creator = users[card.creator]
        card.labels = ', '.join([labels.get(lab, lab) for lab in lab.split(', ')])
    return cards

def filter_cards(cards,cards_projects=[]):
    '''Фильтрует карточки по названиям проектов'''
    for card in cards:
        for label in card.labels.split(', '):
             if label in TIME_LABELS:
                card.hours += TIME_LABELS[label]
             if label in LABELS:
                card.labels = label
                cards_projects.append(card)
    return cards_projects


def result_table(cards_projects, results=[], creators=[]):
    '''Создает объект с количеством часов для каждого проекта'''
    project_list=[]
    creators_list=[]
    for card in cards_projects:
        result = Result(project=card.labels,
                        hours=card.hours,
                        )
        if result.project not in project_list:
            results.append(result)
            project_list.append(result.project)
        elif result.project in project_list:
            for result in results:
                if result.project == card.labels:
                    result.hours += card.hours

        creator = Creator(creator=card.creator,
                          creator_hours=card.hours
                          )
        if creator.creator not in creators_list:
            creators.append(creator)
            creators_list.append(creator.creator)
        elif creator.creator in creators_list:
            for creator in creators:
                if creator.creator == card.creator:
                    creator.creator_hours += card.hours

    return results, creators



def create_xslx(path, cards_projects, results, creators):
    '''Создает таблицы эксель'''
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Проект', 'Создатель', 'Описание', 'Часы'])
    for card in cards_projects:
        data = [card.labels, card.creator, card.title, card.hours]
        sheet.append(data)
    set_style_cards(sheet)
    workbook_res = Workbook()
    sheet_res = workbook_res.active
    sheet_res.append(['Проект','Часы'])
    for result in results:
        data = [result.project,result.hours]
        sheet_res.append(data)
    sheet_res.append([''])
    sheet_res.append(['Сотрудник','Часы'])
    n = len(sheet_res['A'])
    for creator in creators:
        data = [creator.creator, creator.creator_hours]
        sheet_res.append(data)
    set_style_creators(sheet_res, n)


    workbook.save(filename=f"{path}/Cards.xlsx")
    workbook_res.save(filename=f"{path}/Projects.xlsx")



def main(input_args=None):
    args = parse_args(input_args)
    input_path = Path(args.input_path)
    output_path = Path(args.output_path)
    assert input_path != output_path
    data = json_reader(input_path)
    cards = get_cards(data)
    cards = change_id_to_name(data, cards)
    cards_projects = filter_cards(cards)

    results, creators = result_table(cards_projects)


    create_xslx(output_path,cards_projects , results, creators)


if __name__ == "__main__":
    main()
