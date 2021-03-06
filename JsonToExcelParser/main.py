import json
import argparse
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

from classes import Card
from config import LABELS, TIME_LABELS, OTHER_LABELS
from styles import set_style_cards, set_style_creators


def parse_args(input_args=None):
    parser = argparse.ArgumentParser(description="Создает файлы excel с необходимыми данными.")
    parser.add_argument("--input_path", help="Входной путь к файлу", required=True)
    parser.add_argument("--output_path", help="Результат", required=True)
    args = parser.parse_known_args(input_args)[0]
    return args


def json_reader(filename, data=None):
    '''Открывает json-файл и конвертирует его в объект python'''
    with open(filename, "r") as file:
        data = json.load(file)
    return data


def get_cards(data):
    '''Создает список объектов Card из json-файла '''
    cards = []
    for card in data["cards"]:
        if card["assignees"] != []:
            card = Card(id=card["_id"],
                        creator=card["assignees"][0],
                        labels=(",".join(card["labelIds"])),
                        title=card["title"])
        else:
            card = Card(id=card["_id"],
                        labels=(",".join(card["labelIds"])),
                        title=card["title"])
        cards.append(card)
    return cards


def change_id_to_name(data, cards):
    '''Изменяет все id объектов на имена'''
    users = {}
    labels = {}
    for user in data["users"]:
        users[user["_id"]] = user["username"]
    for label in data["labels"]:
        labels[label["_id"]] = label["name"]
    for card in cards:
        lab = card.labels
        if card.creator == "Никто не назначен":
            pass
        else:
            card.creator = users[card.creator]
        card.labels = ",".join([labels.get(lab, lab) for lab in lab.split(",")])
    return cards


def filter_cards(cards):
    '''Фильтрует карточки по названиям проектов'''
    cards_projects = []
    other_projects = []

    for card in cards:
        labels = card.labels.split(",")
        for label in labels:
            if label in OTHER_LABELS:
                other_projects.append(card)
            if label in TIME_LABELS:
                card.hours += TIME_LABELS[label]
            if label in LABELS:
                card.labels = label
                cards_projects.append(card)

    for card in list(other_projects):
        labels = card.labels.split(",")
        for label in labels:
            if label in LABELS:
                other_projects.remove(card)
            if label in OTHER_LABELS:
                card.labels = label
    return cards_projects, other_projects


def create_excel_file(cards, other_cards, path):
    '''Создаёт таблицы excel'''
    in_proj = Card(labels="InProjects",
                   id='',
                   title='',
                   creator='')
    for card in other_cards:
        in_proj.hours += card.hours

    projects_table = pd.DataFrame(cards)
    projects_sum = projects_table.groupby("labels")
    projects_sum = projects_sum.apply(lambda x: x.groupby("id").hours.first().sum())
    projects_sum = projects_sum.reset_index(name="hours")
    in_proj_row = pd.DataFrame({"labels": [in_proj.labels], "hours": [in_proj.hours]})
    projects_sum = projects_sum.append(in_proj_row)
    projects_sum.to_excel(f"{path}/Projects.xlsx", index=False)
    workbook = load_workbook(filename=f"{path}/Projects.xlsx")
    sheet = workbook.active
    set_style_creators(sheet)
    workbook.save(filename=f"{path}/Projects.xlsx")

    for card in other_cards:
        cards.append(card)

    cards_table = pd.DataFrame(cards)
    cards_table.to_excel(f"{path}/Cards.xlsx", index=False)
    workbook = load_workbook(filename=f"{path}/Cards.xlsx")
    sheet = workbook.active
    set_style_cards(sheet)
    workbook.save(filename=f"{path}/Cards.xlsx")

    employee_table = cards_table.groupby("creator")
    employee_table = employee_table.apply(lambda x: x.groupby("id").hours.first().sum())
    employee_table = employee_table.reset_index(name="hours")
    employee_table.to_excel(f"{path}/Employee.xlsx", index=False)
    workbook = load_workbook(filename=f"{path}/Employee.xlsx")
    sheet = workbook.active
    set_style_creators(sheet, employee=True)
    workbook.save(filename=f"{path}/Employee.xlsx")


def main(input_args=None):
    args = parse_args(input_args)
    input_path = Path(args.input_path)
    output_path = Path(args.output_path)
    assert input_path != output_path
    data = json_reader(input_path)
    cards = get_cards(data)
    cards = change_id_to_name(data, cards)
    cards_projects, other_projects = filter_cards(cards)
    create_excel_file(cards_projects, other_projects, output_path)


if __name__ == "__main__":
    main()
